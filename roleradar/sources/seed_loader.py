"""Load local taxonomy seed files into the database."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from sqlalchemy.orm import Session

from roleradar.storage.repositories import SkillRepository


@dataclass(frozen=True)
class TaxonomySeedRow:
    """One skill taxonomy seed row."""

    skill_name: str
    category: str | None
    source_taxonomy: str
    alias: str
    match_type: str
    case_sensitive: bool


@dataclass(frozen=True)
class TaxonomyLoadResult:
    """Summary of a taxonomy load operation."""

    rows_read: int
    skills_seen: int
    aliases_seen: int


def load_taxonomy_seed(session: Session, file_path: str | Path) -> TaxonomyLoadResult:
    """Load skills and aliases from a CSV or XLSX seed file."""
    rows = list(read_taxonomy_seed(file_path))
    repository = SkillRepository(session)
    skill_keys: set[tuple[str, str]] = set()
    aliases_seen = 0

    for row in rows:
        skill = repository.get_or_create_skill(
            name=row.skill_name,
            category=row.category,
            source_taxonomy=row.source_taxonomy,
        )
        repository.get_or_create_alias(
            skill=skill,
            alias=row.alias,
            match_type=row.match_type,
            case_sensitive=row.case_sensitive,
        )
        skill_keys.add((row.skill_name.casefold().strip(), row.source_taxonomy))
        aliases_seen += 1

    session.flush()
    return TaxonomyLoadResult(
        rows_read=len(rows),
        skills_seen=len(skill_keys),
        aliases_seen=aliases_seen,
    )


def read_taxonomy_seed(file_path: str | Path) -> Iterable[TaxonomySeedRow]:
    """Read taxonomy rows from a supported seed file."""
    path = Path(file_path)
    if path.suffix.casefold() == ".csv":
        yield from _read_csv_seed(path)
        return

    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        yield from _read_xlsx_seed(path)
        return

    raise ValueError(f"Unsupported taxonomy seed format: {path.suffix}")


def _read_csv_seed(path: Path) -> Iterable[TaxonomySeedRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for raw_row in reader:
            yield _parse_row(raw_row)


def _read_xlsx_seed(path: Path) -> Iterable[TaxonomySeedRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading Excel taxonomy seeds requires openpyxl to be installed."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]

    for values in rows:
        raw_row = {
            header[index]: value
            for index, value in enumerate(values)
            if index < len(header)
        }
        yield _parse_row(raw_row)


def _parse_row(raw_row: dict[str, object]) -> TaxonomySeedRow:
    skill_name = _clean(raw_row.get("skill_name"))
    if not skill_name:
        raise ValueError("Taxonomy seed row is missing skill_name")

    alias = _clean(raw_row.get("alias")) or skill_name

    return TaxonomySeedRow(
        skill_name=skill_name,
        category=_clean(raw_row.get("category")) or None,
        source_taxonomy=_clean(raw_row.get("source_taxonomy")) or "local",
        alias=alias,
        match_type=_clean(raw_row.get("match_type")) or "literal",
        case_sensitive=_parse_bool(raw_row.get("case_sensitive")),
    )


def _clean(value: object) -> str:
    return str(value or "").strip()


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in {"1", "true", "yes", "y"}

